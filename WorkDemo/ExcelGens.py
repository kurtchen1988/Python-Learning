import pymysql, time
import openpyxl



class excelGen:

    sql1 = "select org_code from db_user.vanyar_org_institution where org_code = '%s';" # 组织机构代码
    sql2 = "select name from db_user.vanyar_org_institution where name = '%s' and `status` = 1;" # 机动车所有人
    sql3 = "select plate_no from db_car_manage.cm_car_stock where plate_no = '%s';" # 车辆号
    sql4 = "SELECT name FROM db_car_manage.cm_car_class WHERE district_code='330000' AND status='0' and name='%s';" #车辆类型
    sql5 = "SELECT name FROM db_car_manage.cm_car_usage WHERE district_code='330000' AND status='0' and name='%s';" #车辆性质
    sql6 = "SELECT name FROM db_item.parana_brands where  name = '%s';" #车辆品牌：
    sql7 = "select plate_no from  db_car_manage.cm_car_stock where vin_no = '%s';" #车架号
    sql8 = "select plate_no from  db_car_manage.cm_car_stock where engine_no = '%s';" #发动机号

    user = 'devread'
    passwd = 'devR3343'
    urlprod = 'rr-9dpnbsn9mds4urq9g.mysql.rds.aliyuncs.com'
    urlitems = 'rr-9dp30iap52gda17cz.mysql.rds.aliyuncs.com'
    urluser = 'rr-9dp8q4d5d7chv3uy2.mysql.rds.aliyuncs.com'

    goodData = []
    badData = []

    failMessage = []

    fileExcelGood = openpyxl.Workbook()
    fileExcelBad = openpyxl.Workbook()

    def __init__(self):
        self.dbprd = pymysql.connect(host=self.urlprod, user=self.user, password=self.passwd, port=3306,
                                     charset='utf8')  # production数据库
        self.dbitem = pymysql.connect(host=self.urlitems, user=self.user, password=self.passwd, port=3306,
                                      charset='utf8')  # item数据库
        self.dbuser = pymysql.connect(host=self.urluser, user=self.user, password=self.passwd, port=3306,
                                      charset='utf8')  # user数据库

        self.curprd = self.dbprd.cursor()
        self.curitem = self.dbitem.cursor()
        self.curuser = self.dbuser.cursor()

        self.sheetSuccess = self.fileExcelGood.create_sheet("Success", 0)
        self.sheetFail = self.fileExcelBad.create_sheet("Fail", 0)

    def deterSQL(self, sql, db):
        db.execute(sql)
        res = db.fetchall()
        if res == ():
            return None
        else:
            return res[0][0]



    def readExcel(self, file, row):

        wb = openpyxl.load_workbook(filename=file)
        sheet = wb.active
        content = []

        for col in range(1, 13):
            content.append(sheet.cell(row=row, column=col).value)

        return content


    def writeFile(self,flag, data,row):


        if flag == 1:
            col = 1
            for val in data:
                self.sheetSuccess.cell(row=row, column=col).value = val
                col = col + 1
        elif flag == 0:
            col = 1
            for val in data:
                self.sheetFail.cell(row=row, column=col).value = val
                col = col + 1
        #fileExcel.save(file)

    def controlMain(self, file):
        wb = openpyxl.load_workbook(filename=file)
        file1 = "成功数据" + str(time.time()) + '.xlsx'
        file2 = "失败数据" + str(time.time()) + '.xlsx'
        rowFile1 = 2
        rowFile2 = 2
        sheetOri = wb.active

        oriRow = sheetOri.max_row

        for row in range(1, oriRow + 1):
            if row == 1:
                data = self.readExcel(file,row)
                self.writeFile(1,data, row)
                data.append("数据验证失败原因")
                self.writeFile(0,data, row)
            else:
                data = self.readExcel(file, row)
                if self.processData(data):
                    self.writeFile(1,data,rowFile1)
                    rowFile1=rowFile1+1
                else:
                    more = ','.join(self.failMessage)
                    data.append(more)
                    self.writeFile(0,data,rowFile2)
                    self.failMessage=[]
                    rowFile2=rowFile2+1

        self.fileExcelGood.save(file1)
        self.fileExcelBad.save(file2)


    def processData(self, data):
        flag = []


        # 机构代码,单位名称
        if self.deterSQL(self.sql1%data[0], self.curuser)!=None and self.deterSQL(self.sql2%data[1], self.curuser)!=None:
            flag.append(0)
        if self.deterSQL(self.sql1%data[0], self.curuser)==None:
            self.failMessage.append("机构代码不一致")
            #print("机构代码不一致")
            flag.append(1)
        if self.deterSQL(self.sql2%data[1], self.curuser)==None:
            self.failMessage.append("单位名称不一致")
            #print("单位名称不一致")
            flag.append(1)
        if self.deterSQL(self.sql1 % data[0], self.curuser) == None and self.deterSQL(self.sql2 % data[1], self.curuser) == None:
            self.failMessage.append("单位不存在")
            #print("单位不存在")
            flag.append(1)

        # 查询车牌号
        if self.deterSQL(self.sql3%data[2], self.curprd)!=None:
            self.failMessage.append("车牌号重复")
            #print("车牌号重复")
            flag.append(1)
        if self.deterSQL(self.sql3%data[2], self.curprd)==None:
            flag.append(0)

        # 查询车辆类型
        if self.deterSQL(self.sql4%data[3], self.curprd)!=None:

            flag.append(0)
        if self.deterSQL(self.sql4%data[3], self.curprd)==None:
            self.failMessage.append("车辆类型不存在")
            #print("车辆类型不存在")
            flag.append(1)

        # 查询车辆性质
        if self.deterSQL(self.sql5%data[4], self.curprd)!=None:

            flag.append(0)
        if self.deterSQL(self.sql5%data[4], self.curprd)==None:
            self.failMessage.append("车辆性质不存在")
            #print("车辆性质不存在")
            flag.append(1)

        # 查询车辆品牌
        if self.deterSQL(self.sql6 % data[7], self.curitem) != None:
            flag.append(0)
        if self.deterSQL(self.sql6 % data[7], self.curitem) == None:
            self.failMessage.append("车辆品牌不存在")
            #print("车辆品牌不存在")
            flag.append(1)

        # 查询车架号
        if self.deterSQL(self.sql7 % data[9], self.curprd) != None:
            self.failMessage.append("车架号重复")
            #print("车架号重复")
            flag.append(1)
        if self.deterSQL(self.sql7 % data[9], self.curprd) == None:
            flag.append(0)

        # 查询发动机号
        if self.deterSQL(self.sql8 % data[10], self.curprd) != None:
            self.failMessage.append("发动机号重复"+self.deterSQL(self.sql8 % data[10], self.curprd))
            #print("发动机号重复")
            flag.append(1)
        if self.deterSQL(self.sql8 % data[10], self.curprd) == None:
            flag.append(0)


        #print(flag)
        #print(self.failMessage)

        s = 0
        for x in flag:
            s += x

        if s == 0:
            return True
        else:
            return False


    def __del__(self):

        self.curprd.close()
        self.curitem.close()
        self.curuser.close()

        self.dbprd.close()
        self.dbitem.close()
        self.dbuser.close()

if __name__ == '__main__':

    a = excelGen()
    #print(a.readExcel('兰溪市森林公安局.xlsx',2))
    #a.writeFile("test.xlsx",[1,2,3],1)
    #print(a.processData(a.readExcel('兰溪市森林公安局.xlsx',2)))
    print("使用说明：将本Exe文件放在任意文件夹中，然后输入你需要整理的xlsx文件的文件名并按回车。\n"
          "会有两个带有时间戳的xlsx文件在同文件夹生成，分别是成功验证的数据和失败验证的数据。\n"
          "暂时只支持单个文件验证...")
    keyin = input("请输入需要处理的文件名称并回车：")
    a.controlMain(keyin+'.xlsx')