import pymysql, time,datetime,sys
import openpyxl, re



class excelGen:

    sql1 = "select org_code from db_user.vanyar_org_institution where org_code = '%s';" # 组织机构代码
    sql2 = "select name from db_user.vanyar_org_institution where name = '%s' and `status` = 1;" # 机动车所有人
    sql21 = "select org_code from db_user.vanyar_org_institution where name = '%s' and `status` = 1;" # 查找正确机构代码
    sql3 = "select org_name from db_car_manage.cm_car_stock where plate_no = '%s';" # 车辆号
    sql4 = "SELECT name FROM db_car_manage.cm_car_class WHERE district_code='330000' AND status='0' and name='%s';" #车辆类型
    sql5 = "SELECT name FROM db_car_manage.cm_car_usage WHERE district_code='330000' AND status='0' and id<>10021 and name='%s';" #车辆性质
    sql6 = "SELECT name FROM db_item.parana_brands where  name like '%%%s%%';" #车辆品牌：
    sql7 = "select plate_no from  db_car_manage.cm_car_stock where vin_no = '%s';" #车架号
    sql8 = "select plate_no from  db_car_manage.cm_car_stock where engine_no = '%s';" #发动机号
    sql9 = "select org_name from  db_car_manage.cm_car_stock where vin_no = '%s';" #车架号
    sql10 = "select org_name from  db_car_manage.cm_car_stock where engine_no = '%s';" #发动机号

    user = 'devread'
    passwd = 'devR3343'
    urlprod = 'rr-9dpnbsn9mds4urq9g.mysql.rds.aliyuncs.com'
    urlitems = 'rr-9dp30iap52gda17cz.mysql.rds.aliyuncs.com'
    urluser = 'rr-9dp8q4d5d7chv3uy2.mysql.rds.aliyuncs.com'

    goodData = [] # 好数据数组
    badData = [] # 不好数据数组
    validate = [None] * 6 # 验证修改数据数组，默认长度为7，分别是要验证的数据


    failMessage = ['error']
    fileName = '默认'



    fileExcelGood = openpyxl.Workbook()
    fileExcelBad = openpyxl.Workbook()

    def __init__(self):
        self.dbprd = pymysql.connect(host=self.urlprod, user=self.user, password=self.passwd, port=3306,
                                     charset='utf8')  # production数据库
        self.dbitem = pymysql.connect(host=self.urlitems, user=self.user, password=self.passwd, port=3306,
                                      charset='utf8')  # item数据库
        self.dbuser = pymysql.connect(host=self.urluser, user=self.user, password=self.passwd, port=3306,
                                      charset='utf8')  # user数据库

        self.curprd = self.dbprd.cursor()
        self.curitem = self.dbitem.cursor()
        self.curuser = self.dbuser.cursor()

        self.sheetSuccess = self.fileExcelGood.create_sheet("Success", 0)
        self.sheetFail = self.fileExcelBad.create_sheet("Fail", 0)

    def deterSQL(self, sql, db):
        db.execute(sql)
        res = db.fetchall()
        if res == ():
            return None
        else:
            return res[0][0]



    def readExcel(self, file, row):

        wb = openpyxl.load_workbook(filename=file)
        sheet = wb.active
        content = []

        for col in range(1, 13):
            if type(sheet.cell(row=row, column=col).value)=='str':
                content.append(str.strip(sheet.cell(row=row, column=col).value))
            else:
                content.append(sheet.cell(row=row, column=col).value)

        return content


    def writeFile(self,flag, data,row):
        '''

        :param flag: 1是往成功的excel中写，2是往失败的excel中写
        :param data: 写入的数据
        :param row: 写入的行数
        :return:
        '''

        if flag == 1:
            col = 1
            for val in data:
                if col == 4 and self.validate[0]!=None: #
                    self.sheetSuccess.cell(row=row, column=col).value = self.validate[0]

                elif col == 5 and self.validate[1] != None: #
                    self.sheetSuccess.cell(row=row, column=col).value = self.validate[1]

                elif col == 6 and self.validate[2] != None: #
                    self.sheetSuccess.cell(row=row, column=col).value = self.validate[2]

                elif col == 7 and self.validate[3] != None: #
                    self.sheetSuccess.cell(row=row, column=col).value = self.validate[3]

                elif col == 9 and self.validate[4] != None: #
                    self.sheetSuccess.cell(row=row, column=col).value = self.validate[4]

                elif col == 12 and self.validate[5] != None: #
                    self.sheetSuccess.cell(row=row, column=col).value = self.validate[5]

                else: # 其它
                    self.sheetSuccess.cell(row=row, column=col).value = val
                #self.sheetSuccess.cell(row=row, column=col).value = val
                col = col + 1
              

        elif flag == 0:
            col = 1
            for val in data:
                if col == 4 and self.validate[0]!=None: #
                    self.sheetFail.cell(row=row, column=col).value = self.validate[0]

                elif col == 5 and self.validate[1] != None: #
                    self.sheetFail.cell(row=row, column=col).value = self.validate[1]

                elif col == 6 and self.validate[2] != None: #
                    self.sheetFail.cell(row=row, column=col).value = self.validate[2]

                elif col == 7 and self.validate[3] != None: #
                    self.sheetFail.cell(row=row, column=col).value = self.validate[3]

                #elif col == 8 and self.validate[6] != None: #
                    #self.sheetFail.cell(row=row, column=col).value = self.validate[6]

                elif col == 9 and self.validate[4] != None: #
                    self.sheetFail.cell(row=row, column=col).value = self.validate[4]

                elif col == 12 and self.validate[5] != None: #
                    self.sheetFail.cell(row=row, column=col).value = self.validate[5]

                else: # 其它
                    self.sheetFail.cell(row=row, column=col).value = val
                col = col + 1
               
        #fileExcel.save(file)

    def modifyData(self):
        pass


    def controlMain(self, file):
        wb = openpyxl.load_workbook(filename=file)
        file1 = "success_" + self.fileName + str(datetime.datetime.now().strftime("%Y%m%d%H%M%S")) + '.xlsx'
        file2 = "failure_" + self.fileName + str(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))+ '.xlsx'
        rowFile1 = 2
        rowFile2 = 2
        sheetOri = wb.active

        oriRow = sheetOri.max_row

        print("一共"+str(oriRow)+"条数据")
        output = sys.stdout

        for row in range(1, oriRow + 1):
            output.write(str('%.2f%%'%((row/oriRow)*100)))

            if row == 1:
                data = self.readExcel(file,row)
                self.writeFile(1,data, row)
                data.append("失败原因")
                self.writeFile(0,data, row)
            else:
                data = self.readExcel(file, row)
                if self.processData(data):
                    self.writeFile(1,data,rowFile1)
                    rowFile1=rowFile1+1
                else:
                    more = ','.join(self.failMessage)
                    data.append(more)
                    self.writeFile(0,data,rowFile2)
                    # self.failMessage=[]
                    self.validate = [None] * 6
                    rowFile2=rowFile2+1

            self.failMessage = []
        sys.stdout.flush()
        self.fileExcelGood.save(file1)
        self.fileExcelBad.save(file2)


    def processData(self, data):
        flag = []
        carType1 = ["应急机要公务用车","调研用车","一般工作用车","其他用车","机关领导用车","接待用车","一般公务用车","其他","生产经营性用车","公务用车"]
        carType2 = ["老干部服务用车","老干部用车"]
        carType3 = ["一般执法执勤用车","执法执勤用车","一般执勤执法用车"]
        carType4 = ["其他特种专业技术用车","其他特种车辆","特殊专用技术用车","非营业机关、事业团体客车","其他特种执法用车","非营运车辆","特种用车","特种专业技术用车","其他特种专业技术用车"]

        # 机构代码,单位名称
        if self.deterSQL(self.sql1%data[0], self.curuser)!=None and self.deterSQL(self.sql2%data[1], self.curuser)!=None:
            flag.append(0)
        if self.deterSQL(self.sql1%data[0], self.curuser)==None:
            self.failMessage.append("机构代码不一致")
            #print("机构代码不一致")
            flag.append(1)
        if self.deterSQL(self.sql2%data[1], self.curuser)==None:
            self.failMessage.append("单位名称不一致")
            #print("单位名称不一致")
            flag.append(1)
        # 增加的对组织代码修改
        if self.deterSQL(self.sql2%data[1], self.curuser)!=None:
            data[0] = self.deterSQL(self.sql21%data[1], self.curuser)
        if self.deterSQL(self.sql1 % data[0], self.curuser) == None and self.deterSQL(self.sql2 % data[1], self.curuser) == None:
            self.failMessage.append("单位平台不存在")
            #print("单位不存在")
            flag.append(1)

        # 查询车牌号
        if self.deterSQL(self.sql3%data[2], self.curprd)!=None:
            if (self.deterSQL(self.sql3%data[2], self.curprd)==data[1]):
                self.failMessage.append(str(data[2]) + "-车牌号重复(已在该单位下)")
            else:
                self.failMessage.append(str(data[2])+"-车牌号重复(归属单位："+self.deterSQL(self.sql3%data[2], self.curprd)+")")
            #print("车牌号重复")
            flag.append(1)
        if self.deterSQL(self.sql3%data[2], self.curprd)==None:
            flag.append(0)

        # 判断车辆类型正确否并存入列表
        if data[3] == '小汽车' or data[3] == '小轿车':
            self.validate[0] = '轿车'
        elif data[3] == '小型客车':
            self.validate[0] = '轻型客车'
        elif data[3] == '面包车':
            self.validate[0] = '商务面包车'
        elif data[3] == '工具车':
            self.validate[0] = '特种改装车'


        # 查询车辆类型
        if self.deterSQL(self.sql4%data[3], self.curprd)!=None:

            flag.append(0)
        if self.deterSQL(self.sql4%data[3], self.curprd)==None:
            self.failMessage.append(str(data[3])+"-车辆类型不存在")
            #print("车辆类型不存在")
            flag.append(1)

        # 判断车辆性质并存入列表
        if data[4] == '机要通信用车':
            self.validate[1]='机要通信用车'
            self.failMessage.append("车辆性质原值为：机要通信用车")
        elif data[4] == '省部级领导用车':
            self.validate[1] = '省部级领导用车'
            self.failMessage.append("车辆性质原值为：省部级领导用车")
        elif data[4] == '行政执法专用车':
            self.validate[1] = '行政执法专用车'
            self.failMessage.append("车辆性质原值为：行政执法专用车")
        elif data[4] in carType1:
            self.validate[1] = '应急保障用车'
            self.failMessage.append("车辆性质原值为："+ str(data[4]))
        elif data[4] in carType2:
            self.validate[1] = '老干部服务用车'
            self.failMessage.append("车辆性质原值为：" + str(data[4]))
        elif data[4] in carType3:
            self.validate[1] = '执法执勤用车'
            self.failMessage.append("车辆性质原值为：" + str(data[4]))
        elif data[4] in carType4:
            self.validate[1] = '其他特种专业技术用车'
            self.failMessage.append("车辆性质原值为：" + str(data[4]))


        # 查询车辆性质
        if self.deterSQL(self.sql5%data[4], self.curprd)!=None:

            flag.append(0)
        if self.deterSQL(self.sql5%data[4], self.curprd)==None:
            self.failMessage.append(str(data[4])+"-车辆性质不存在")
            #print("车辆性质不存在")
            flag.append(1)

        # 车价格式整合
        if re.match('\d*\.*\d*万',str(data[5])):
            num = re.findall('(\d*\.*\d*)万', str(data[5]))[0]
            num = float(num)
            num = round(num,2)
            self.validate[2] = num
            self.failMessage.append("车辆价格原值为：" + str(data[5]))
        elif re.match('\d*\.*\d*元', str(data[5])):
            num = re.findall('(\d*\.*\d*)元', str(data[5]))[0]
            num = float(num)
            num = round(num,2)
            self.validate[2] = num
            self.failMessage.append("车辆价格原值为：" + str(data[5]))
        elif re.match('\d*\.*\d{1,}',str(data[5])):
            num = re.findall('\d*\.*\d*', str(data[5]))[0]
            num = float(num)
            num = round(num,2)
            self.validate[2] = num
            self.failMessage.append("车辆价格原值为：" + str(data[5]))
        elif data[5] == None or data[5] == '':
            self.validate[2] = 1
            self.failMessage.append("车辆价格原值为空")

        # 登记日期整理
        if re.match('\d{4}.\d{1,2}.?$', str(data[6])):
            whole = re.findall('(\d{4}).(\d{1,2}).*$', str(data[6]))
            year = whole[0][0]
            month = int(whole[0][1])
            month = str(month)
            self.validate[3] = year+'-'+month+'-'+'1'
            self.failMessage.append("车辆原登记日期为："+str(data[6]))

        elif re.match('\d{4}.\d{1,2}.\d{1,2}.*$', str(data[6])):
            whole = re.findall('(\d{4}).(\d{1,2}).(\d{1,2}).*$', str(data[6]))
            year = whole[0][0]
            month = int(whole[0][1])
            month = str(month)
            day = int(whole[0][2])
            day = str(day)
            self.validate[3] = year+'-'+month+'-'+day
            self.failMessage.append("车辆原登记日期为："+str(data[6]))

        # 查询车辆品牌
        if self.deterSQL(self.sql6 % data[7], self.curitem) != None:
            flag.append(0)
        if self.deterSQL(self.sql6 % data[7], self.curitem) == None:
            self.failMessage.append(str(data[7])+"-车辆品牌不存在")
            #print("车辆品牌不存在")
            flag.append(1)

        # （后加）整理品牌
        data[7] = str.strip(str(data[7]))

        if re.match('.*牌.*', str(data[7])):
            #brand = str.strip(str(data[7]))
            brand = str(data[7]).strip('牌')
            data[7] = brand
            #self.failMessage.append("车辆品牌原为：" + str(data[7]))

        # 整理排量
        if re.match('\d{4,}\w*', str(data[8])):
            first = re.findall('(\d{4,})\w*', str(data[8]))
            first = first[0]
            first = float(first)
            first = round(first / 1000, 1)
            self.validate[4]=first
            self.failMessage.append("车辆排量原为：" + str(data[8]))

        elif re.match('\d{1,2}\.\d*T', str(data[8])):
            second = re.findall('(\d{1,2}\.\d*)T', str(data[8]))
            second = second[0]
            second = float(second)
            second = round(second, 1)
            self.validate[4]=second
            self.failMessage.append("车辆排量原为：" + str(data[8]))

        elif re.match('\d{1,2}\.\d{2,}', str(data[8])):
            third = re.findall('\d{1,2}\.\d{2,}', str(data[8]))
            third = third[0]
            third = float(third)
            third = round(third, 1)
            self.validate[4]=third
            self.failMessage.append("车辆排量原为：" + str(data[8]))
        elif re.match('\d{1,2}\.\d{1,}\D',str(data[8])):
            forth = re.findall('(\d{1,2}\.\d{1,})\D',str(data[8]))
            forth = forth[0][0]
            forth = float(forth)
            forth = round(forth,1)
            self.validate[4] = forth
            self.failMessage.append("车辆排量原为：" + str(data[8]))
        elif data[8] == None or data[8] == '':
            self.validate[4] = 0
            self.failMessage.append("车辆排量原为空")

        # 查询车架号
        if self.deterSQL(self.sql7 % data[9], self.curprd) != None:
            if(self.deterSQL(self.sql7 % data[9], self.curprd)==data[2]):
                self.failMessage.append(str(data[2])+":车架号重复(属同一车辆/平台已存在)")
            else:
                self.failMessage.append(str(data[2])+"的车架号["+data[9]+"]重复(已被["+self.deterSQL(self.sql9 % data[9], self.curprd)+"-"+self.deterSQL(self.sql7 % data[9], self.curprd)+"]占用，请提供行驶证的完整车架号)")
            #print("车架号重复")
            flag.append(1)
        if self.deterSQL(self.sql7 % data[9], self.curprd) == None:
            flag.append(0)

        # 查询发动机号
        if self.deterSQL(self.sql8 % data[10], self.curprd) != None:
            if (self.deterSQL(self.sql8 % data[10], self.curprd) == data[2]):
                self.failMessage.append(str(data[2])+":发动机号重复(属同一车辆/平台已存在)")
            else:
                self.failMessage.append(str(data[2])+"的发动机号["+data[10]+"]重复(已被["+self.deterSQL(self.sql10 % data[10], self.curprd)+"-"+self.deterSQL(self.sql8 % data[10], self.curprd)+"]占用，请提供行驶证的完整发动机号)")
            #print("发动机号重复")
            flag.append(1)
        if self.deterSQL(self.sql8 % data[10], self.curprd) == None:
            flag.append(0)



        # 整理进口
        if data[11]=="是":
            self.validate[5] = "进口"
            self.failMessage.append("车辆进口原为：" + str(data[8]))
        elif data[11]=='否':
            self.validate[5] = "国产"
            self.failMessage.append("车辆进口原为：" + str(data[8]))



        s = 0
        for x in flag:
            s += x

        if s == 0:
            return True
        else:
            return False

    def setFileName(self, fileName):
        self.fileName = fileName

    def __del__(self):

        self.curprd.close()
        self.curitem.close()
        self.curuser.close()

        self.dbprd.close()
        self.dbitem.close()
        self.dbuser.close()

if __name__ == '__main__':

    a = excelGen()
    print("说明：请将要处理的xlsx文件放在此程序同目录下，暂时只支持一次处理一个xlsx文件。")
    keyin = input("请输入需要处理的文件名称并回车(不用加xlsx后缀)：")
    a.setFileName(keyin)
    a.controlMain("./"+keyin+'.xlsx')
    # a.controlMain('./123.xlsx')
